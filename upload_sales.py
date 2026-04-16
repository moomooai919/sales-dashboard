#!/usr/bin/env python3
"""
Steam CDKey 銷售數據上傳工具
用法：python upload_sales.py <Excel文件路徑>
"""
import sys, json, base64, urllib.request, subprocess
from pathlib import Path
from datetime import datetime
from collections import defaultdict

GITHUB_TOKEN = "YOUR_GITHUB_TOKEN"   # ← 替換成你的 token
REPO = "moomooai919/sales-dashboard"  # ← 你的 public repo
DATA_FILE = "data/sales.json"

def parse_excel(path):
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
    
    wb = openpyxl.load_workbook(path, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [str(c.value or '').strip() for c in ws[1]]
        if any('订单' in h or 'Order' in h for h in headers):
            break
    
    def ci(names):
        for n in names:
            for i, h in enumerate(headers):
                if n.lower() in h.lower(): return i
        return None
    
    idx = {
        'id': ci(['订单ID','Order ID']),
        'time': ci(['下单时间','Order Time']),
        'prod': ci(['游戏','Product']),
        'sales': ci(['销售额','Sales']),
        'inc': ci(['收入','Income']),
        'status': ci(['交易状态','Status']),
    }
    
    orders = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        oid = str(row[idx['id']] or '').strip()
        if not oid: continue
        tv = row[idx['time']]
        if isinstance(tv, datetime): dt = tv
        else:
            try: dt = datetime.strptime(str(tv)[:19], '%Y-%m-%d %H:%M:%S')
            except: continue
        orders[oid] = {
            'order_id': oid,
            'date': dt.strftime('%Y-%m-%d'),
            'month': dt.strftime('%Y-%m'),
            'product': str(row[idx['prod']] or ''),
            'sales': float(row[idx['sales']] or 0),
            'income': float(row[idx['inc']] or 0),
            'status': str(row[idx['status']] or ''),
            'channel': 'SteamPY',
        }
    return orders

def get_current_data(token, repo, file_path):
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"
    req = urllib.request.Request(url, headers={"Authorization": f"token {token}", "User-Agent": "bot"})
    try:
        resp = urllib.request.urlopen(req, timeout=15)
        data = json.loads(resp.read())
        content = json.loads(base64.b64decode(data['content']))
        return content, data['sha']
    except:
        return None, None

def build_summary(all_orders):
    monthly = defaultdict(lambda: {'orders':0,'sales':0,'income':0})
    products = defaultdict(lambda: {'orders':0,'sales':0,'income':0})
    for o in all_orders.values():
        m = o['month']
        monthly[m]['orders'] += 1
        monthly[m]['sales'] += o['sales']
        monthly[m]['income'] += o['income']
        p = o['product']
        products[p]['orders'] += 1
        products[p]['sales'] += o['sales']
        products[p]['income'] += o['income']
    
    monthly_list = [{'month':m,'orders':v['orders'],'sales':round(v['sales'],2),'income':round(v['income'],2)} for m,v in sorted(monthly.items())]
    products_list = [{'product':p,'orders':v['orders'],'sales':round(v['sales'],2),'income':round(v['income'],2)} for p,v in sorted(products.items(),key=lambda x:-x[1]['orders'])]
    
    months = sorted(monthly.keys())
    mom = 0
    if len(months) >= 2:
        last = monthly[months[-1]]['sales']
        prev = monthly[months[-2]]['sales']
        if prev: mom = round((last-prev)/prev*100, 1)
    
    total_orders = sum(v['orders'] for v in monthly.values())
    total_sales = round(sum(v['sales'] for v in monthly.values()), 2)
    total_income = round(sum(v['income'] for v in monthly.values()), 2)
    latest = max(o['date'] for o in all_orders.values()) if all_orders else ''
    
    return {
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'latest_order_date': latest,
        'total_orders': total_orders,
        'total_sales': total_sales,
        'total_income': total_income,
        'mom_growth': mom,
        'monthly': monthly_list,
        'products': products_list,
    }

def push_json(token, repo, file_path, content, sha=None):
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"
    payload = {
        "message": f"data: update sales {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "content": base64.b64encode(json.dumps(content, ensure_ascii=False, indent=2).encode()).decode(),
    }
    if sha: payload["sha"] = sha
    req = urllib.request.Request(url, data=json.dumps(payload).encode(), method='PUT',
                                  headers={"Authorization": f"token {token}", "Content-Type": "application/json", "User-Agent":"bot"})
    resp = urllib.request.urlopen(req, timeout=20)
    return json.loads(resp.read())

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法：python upload_sales.py <Excel文件路徑>")
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    print(f"📂 解析：{xlsx_path}")
    new_orders = parse_excel(xlsx_path)
    print(f"  找到 {len(new_orders)} 筆訂單")
    
    print("📡 從 GitHub 拉取現有數據…")
    current, sha = get_current_data(GITHUB_TOKEN, REPO, DATA_FILE)
    
    # 合併（保留舊訂單，新訂單覆蓋）
    # 注意：summary 模式只有統計，這裡直接用新文件重算+追加邏輯
    # 如需完整合併，需要也存儲完整明細（可擴展）
    print("🔀 合併數據…")
    summary = build_summary(new_orders)
    
    if current:
        # 將現有月度數據合入（非重疊月份保留）
        existing_months = {m['month']: m for m in current.get('monthly',[])}
        new_months = {m['month']: m for m in summary['monthly']}
        for mo, v in existing_months.items():
            if mo not in new_months:
                summary['monthly'].append(v)
        summary['monthly'].sort(key=lambda x: x['month'])
        
        # 重算 totals
        summary['total_orders'] = sum(m['orders'] for m in summary['monthly'])
        summary['total_sales'] = round(sum(m['sales'] for m in summary['monthly']), 2)
        summary['total_income'] = round(sum(m['income'] for m in summary['monthly']), 2)
    
    print("⬆️  推送到 GitHub…")
    push_json(GITHUB_TOKEN, REPO, DATA_FILE, summary, sha)
    
    print(f"\n✅ 完成！")
    print(f"   總訂單：{summary['total_orders']}")
    print(f"   總銷售額：¥{summary['total_sales']}")
    print(f"   更新時間：{summary['updated_at']}")
    print(f"\n🌐 Dashboard：https://moomooai919.github.io/sales-dashboard/")
